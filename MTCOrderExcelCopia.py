from openpyxl import load_workbook
import requests

class ExelFileCreator:
    def __init__(self, nf_number:int, order_number:int):
        self.nf_number = nf_number
        self.order_number = order_number
    def create(self):
        response = requests.get(
            "https://api.jumpseller.com/v1/orders/status/paid.json?login=ef8cde834d6ae52d628f4655daa78f72&authtoken=f18ce6c94152128475efe4493794f433e8184582a9079d2d4d")

        lst_of_orders = response.json()
        order_number_i = -1
        for i in range(len(lst_of_orders)):
            if lst_of_orders[i]['order']['id'] == self.order_number:
                order_number_i = i
                print(order_number_i)
                break

        for i in range(len(lst_of_orders[order_number_i]['order']['products'])):
            print(lst_of_orders[order_number_i]['order']['products'][i]['sku'],
                  lst_of_orders[order_number_i]['order']['products'][i]['qty'])
        print(order_number_i)
        wb = load_workbook("pedido_template.xlsx")
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[0]]
        for r in range(len(lst_of_orders[order_number_i]['order']['products'])):
            Sheet1.cell(row=r + 2, column=1).value = self.nf_number
            Sheet1.cell(row=r + 2, column=2).value = lst_of_orders[order_number_i]['order']['products'][r]['sku']
            Sheet1.cell(row=r + 2, column=3).value = lst_of_orders[order_number_i]['order']['products'][r]['qty']
            Sheet1.cell(row=r + 2, column=4).value = self.nf_number
            Sheet1.cell(row=r + 2, column=5).value = 1
            Sheet1.cell(row=r + 2, column=6).value = 34554968000181
            Sheet1.cell(row=r + 2, column=7).value = lst_of_orders[order_number_i]['order']['products'][r]['qty']

        wb.save(f'uploads/{self.order_number}_pedido(nf{self.nf_number}).xlsx')

if __name__ == '__main__':
    ExelFileCreator.run(debug=False)
